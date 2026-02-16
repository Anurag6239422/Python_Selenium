'''
Docstring for 30UpdateFile
In this program I learn How to update the file
'''

import openpyxl

def UpdateSheet(file_path, col_val, val, item):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active

    Dict = {}

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row = 1, column=i).value == col_val:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == item:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = val

    book.save(file_path)


    
file_path = r"C:\Users\anura_9posmze\Downloads\download.xlsx"

col_val = input("On which column you want to change the value : ")
val = input("What value you want to change : ")
item = input("Name the Item : ")

UpdateSheet(file_path, col_val, val, item)