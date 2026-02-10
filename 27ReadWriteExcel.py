'''
In this program I learn how to read and write a excel file.
'''

import openpyxl

book = openpyxl.load_workbook(r"C:/Users/anura_9posmze/Downloads/study/Python_Selenium/Python_Selenium/Details.xlsx")

sheet = book.active

cell = sheet.cell(row=1, column=2)

print(cell.value)

sheet.cell(row=3, column=4).value = "Female"

print(sheet.cell(row=3, column=4).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['B3'].value)

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

Dict = {}

for i in range(1,  sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,column=j).value

print(Dict)